"""
Extracts ingredients and nutritional data from every sheet of every .xlsx file
in the Product-nutritional-ingredients-data folder into a single CSV.
"""

import os
import csv
import re
import openpyxl

DATA_DIR = r"C:\Users\Manager\Desktop\Emma Basic Website\Product-nutritional-ingredients-data"
OUTPUT_CSV = r"C:\Users\Manager\Desktop\Emma Basic Website\nutritional_ingredients_master_data.csv"

# Nutritional nutrient labels to capture
NUTRIENT_LABELS = {
    "energy", "fat", "saturate", "monounsaturated", "polyunsaturated",
    "carbohydrate", "sugar", "polyol", "starch", "fibre", "fiber",
    "protein", "salt", "sodium", "vitamin", "mineral", "calcium",
    "iron", "potassium", "phosphorus", "magnesium", "zinc",
}

def clean(val):
    if val is None:
        return ""
    return str(val).replace("\n", " ").strip()

def looks_like_nutrient(label):
    lower = label.lower().strip().lstrip(" *")
    return any(n in lower for n in NUTRIENT_LABELS)

def extract_from_sheet(ws):
    """
    Returns (ingredients_rows, nutrition_rows) from a single worksheet.
    ingredients_rows: list of dicts with keys Ingredient, Components, Composition_pct, Supplier, Country, GMO
    nutrition_rows:   list of dicts with keys Nutrient, Value_per_100g
    """
    ingredients = []
    nutrition = []

    rows = list(ws.iter_rows(values_only=True))

    in_ingredients = False
    in_nutrition = False

    for row in rows:
        cells = [clean(c) for c in row]
        first = cells[0].lower() if cells[0] else ""
        # Use first non-None cell as section marker too
        non_null = [c for c in cells if c]

        # --- Section detection ---
        if any("product composition" in c.lower() for c in non_null):
            in_ingredients = True
            in_nutrition = False
            continue

        if any("nutritional information" in c.lower() for c in non_null):
            in_nutrition = True
            in_ingredients = False
            continue

        if any("fourteen allergen" in c.lower() or "allergen" in c.lower() for c in non_null):
            in_nutrition = False
            in_ingredients = False
            continue

        # --- Skip header/descriptor rows ---
        if any(c.lower().startswith("(including") or c.lower().startswith("(if ingredient") or
               c.lower().startswith("(please") or c.lower().startswith("*please") or
               c.lower().startswith("note:") or c.lower().startswith("are genetically") or
               c.lower().startswith("is gmo") or c.lower().startswith("typical values") for c in non_null):
            in_ingredients_was = in_ingredients
            # After "Typical Values" row, section switch already handled above; keep flag
            if "typical values" in " ".join(non_null).lower():
                pass  # section already set to nutrition above
            continue

        if not non_null:
            continue

        # --- Ingredients section ---
        if in_ingredients:
            # Skip the column header row
            if cells[0].lower() == "ingredients":
                continue
            # Stop at "Total" row or empty ingredient name
            if cells[0] == "" and any("total" in c.lower() for c in cells):
                continue
            ingredient_name = cells[0]
            if not ingredient_name:
                continue
            # col indices: 0=Ingredient, 2=Components, 4=%Composition, 5=Supplier, 6=Country, 7=GMO
            ingredients.append({
                "Ingredient": ingredient_name,
                "Components": cells[2] if len(cells) > 2 else "",
                "Composition_pct": cells[4] if len(cells) > 4 else "",
                "Supplier": cells[5] if len(cells) > 5 else "",
                "Country": cells[6] if len(cells) > 6 else "",
                "GMO": cells[7] if len(cells) > 7 else "",
            })

        # --- Nutrition section ---
        elif in_nutrition:
            nutrient_label = cells[0]
            if not nutrient_label or not looks_like_nutrient(nutrient_label):
                continue
            value = cells[2] if len(cells) > 2 else ""
            nutrition.append({
                "Nutrient": nutrient_label.strip(),
                "Value_per_100g": value,
            })

    return ingredients, nutrition


def find_target_sheet(wb):
    """Returns the sheet containing ingredients/nutrition data."""
    for name in wb.sheetnames:
        lower = name.lower()
        if "ingre" in lower or "nutri" in lower:
            return wb[name], name
    # Fallback: return first sheet
    return wb.active, wb.active.title


def main():
    xlsx_files = sorted(f for f in os.listdir(DATA_DIR) if f.lower().endswith(".xlsx"))
    print(f"Found {len(xlsx_files)} .xlsx files")

    fieldnames = [
        "Source_File", "Sheet_Name", "Data_Type",
        # Ingredient columns
        "Ingredient", "Components", "Composition_pct", "Supplier", "Country", "GMO",
        # Nutrition columns
        "Nutrient", "Value_per_100g",
    ]

    rows_written = 0
    errors = []

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for filename in xlsx_files:
            filepath = os.path.join(DATA_DIR, filename)
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
            except Exception as e:
                errors.append(f"SKIP {filename}: {e}")
                print(f"  ERROR loading {filename}: {e}")
                continue

            # Process every sheet (most files have one relevant sheet, but capture all)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_lower = sheet_name.lower()
                # Only process sheets that contain ingredients/nutrition
                if not any(k in sheet_lower for k in ["ingre", "nutri", "allergen", "composition"]):
                    continue

                try:
                    ingredients, nutrition = extract_from_sheet(ws)
                except Exception as e:
                    errors.append(f"SKIP sheet '{sheet_name}' in {filename}: {e}")
                    print(f"  ERROR in sheet '{sheet_name}' of {filename}: {e}")
                    continue

                for item in ingredients:
                    row = {"Source_File": filename, "Sheet_Name": sheet_name, "Data_Type": "Ingredient"}
                    row.update(item)
                    writer.writerow(row)
                    rows_written += 1

                for item in nutrition:
                    row = {"Source_File": filename, "Sheet_Name": sheet_name, "Data_Type": "Nutrition"}
                    row.update(item)
                    writer.writerow(row)
                    rows_written += 1

            print(f"  OK: {filename}")

    print(f"\nDone. {rows_written} rows written to:\n  {OUTPUT_CSV}")
    if errors:
        print(f"\n{len(errors)} error(s):")
        for e in errors:
            print(f"  - {e}")


if __name__ == "__main__":
    main()
