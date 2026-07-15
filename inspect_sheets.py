import openpyxl, os, sys

sys.stdout.reconfigure(encoding='utf-8')

data_dir = r"C:\Users\Manager\Desktop\Emma Basic Website\Product-nutritional-ingredients-data"
files = [
    'C010 - Dashi Soy Sauce.xlsx',
    'G011 Black Sesame Seed 1kg.xlsx',
    'F103 - Chicken Gyoza 1kg.xlsx',
    'N010 - Soba Noodles.xlsx',
]

for f in files:
    path = os.path.join(data_dir, f)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f'\n=== {f} ===')
    print(f'  Sheets: {wb.sheetnames}')
    # Find the ingredients/nutrition sheet
    for sname in wb.sheetnames:
        if any(k in sname.lower() for k in ['ingre', 'nutri', 'allergen']):
            ws = wb[sname]
            print(f'  --- Sheet: {sname} ---')
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if any(c is not None for c in row):
                    print(f'    Row {i+1}: {[str(c)[:40] if c else None for c in list(row)[:8]]}')
                if i > 80:
                    print('    ...')
                    break
            break
